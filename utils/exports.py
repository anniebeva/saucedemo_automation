import os
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from models.order import Order


def format_order_sheet(sheet):
    """
    Apply formatting to order XLSX sheet
    """

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    info_labels = [
        "Payment Information",
        "Shipping Information",
        "Tax",
        "Total",
    ]

    for row in sheet.iter_rows():
        if row[0].value in info_labels:
            row[0].font = Font(bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    for column_cells in sheet.columns:
        max_length = 0

        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 80)


def export_order_to_xlsx(order: Order, filename: str):
    """
    Export order information to XLSX
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Order"

    sheet.append(
        [
            "Product",
            "Description",
            "Price",
        ]
    )

    for product in order.products:
        sheet.append(
            [
                product.name,
                product.description,
                product.price,
            ]
        )

    sheet.append([])

    sheet.append(["Payment Information", order.payment_information])
    sheet.append(["Shipping Information", order.shipping_information])
    sheet.append(["Tax", order.tax])
    sheet.append(["Total", order.total])

    format_order_sheet(sheet)

    workbook.save(filename)


def wait_pdf(folder, timeout=10):
    """
    Wait until PDF is downloaded
    """

    end = time.time() + timeout

    while time.time() < end:

        for file in os.listdir(folder):
            if file.endswith(".pdf"):
                return file

        time.sleep(0.5)

    raise TimeoutError("PDF was not downloaded.")
